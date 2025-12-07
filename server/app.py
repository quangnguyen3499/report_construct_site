import os
import json
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# -------------------------
# CONFIG
# -------------------------
app = Flask(__name__)
CORS(app, supports_credentials=True)

PORT = int(os.environ.get("PORT", 3001))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
PROJECTS_FILE = os.path.join(DATA_DIR, "projects.json")


# -------------------------
# Ensure Data Directory
# -------------------------
def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(PROJECTS_FILE):
        with open(PROJECTS_FILE, "w") as f:
            json.dump([], f, indent=2)


ensure_data_dir()


# -------------------------
# JSON Helpers
# -------------------------
def read_projects():
    try:
        with open(PROJECTS_FILE, "r") as f:
            return json.load(f)
    except:
        return []


def write_projects(data):
    with open(PROJECTS_FILE, "w") as f:
        json.dump(data, f, indent=2)


# -------------------------
# ROUTES
# -------------------------


@app.get("/api/projects")
def get_projects():
    return jsonify(read_projects())


@app.get("/api/projects/<project_id>")
def get_project(project_id):
    projects = read_projects()
    for p in projects:
        if p["id"] == project_id:
            return jsonify(p)
    return jsonify({"error": "Project not found"}), 404


@app.post("/api/projects")
def create_project():
    projects = read_projects()
    body = request.json

    new_project = {
        "id": str(int(datetime.now().timestamp() * 1000)),
        "name": body.get("name", "New Project"),
        "createdAt": datetime.utcnow().isoformat(),
        "updatedAt": datetime.utcnow().isoformat(),
        "sheets": body.get(
            "sheets",
            {
                "Vật liệu": {"headers": [], "data": []},
                "Nhân công": {"headers": [], "data": []},
                "Máy thi công": {"headers": [], "data": []},
                "Tổng hợp": {"headers": [], "data": []},
            },
        ),
        "data": body.get("data", []),
    }

    projects.append(new_project)
    write_projects(projects)
    return jsonify(new_project)


@app.put("/api/projects/<project_id>")
def update_project(project_id):
    projects = read_projects()
    body = request.json

    for i, p in enumerate(projects):
        if p["id"] == project_id:
            projects[i] = {**p, **body, "updatedAt": datetime.utcnow().isoformat()}
            write_projects(projects)
            return jsonify(projects[i])

    return jsonify({"error": "Project not found"}), 404


@app.delete("/api/projects/<project_id>")
def delete_project(project_id):
    projects = read_projects()
    filtered = [p for p in projects if p["id"] != project_id]
    write_projects(filtered)
    return jsonify({"success": True})


# -------------------------
# STATISTICS
# -------------------------
@app.get("/api/statistics")
def statistics():
    projects = read_projects()

    result = {"totalProjects": len(projects), "projects": []}

    for project in projects:
        materials = {}
        total_cost = 0
        item_count = 0

        # Sheets new structure
        if "sheets" in project:
            for row in project["sheets"].get("Vật liệu", {}).get("data", []):
                name = row.get("tenVatTu")
                qty = float(row.get("khoiLuong", 0))
                cost = float(
                    row.get("thanhTienGiaTB") or row.get("thanhTienGiaGoc") or 0
                )

                if name:
                    if name not in materials:
                        materials[name] = {
                            "name": name,
                            "totalQuantity": 0,
                            "totalCost": 0,
                        }
                    materials[name]["totalQuantity"] += qty
                    materials[name]["totalCost"] += cost
                    total_cost += cost

            item_count = sum(len(s.get("data", [])) for s in project["sheets"].values())

        # Old structure
        if "data" in project and project["data"]:
            for row in project["data"]:
                name = row.get("material")
                qty = float(row.get("quantity", 0))
                unit_price = float(row.get("unitPrice", 0))
                cost = qty * unit_price

                if name:
                    if name not in materials:
                        materials[name] = {
                            "name": name,
                            "totalQuantity": 0,
                            "totalCost": 0,
                        }
                    materials[name]["totalQuantity"] += qty
                    materials[name]["totalCost"] += cost
                    total_cost += cost

            item_count += len(project["data"])

        result["projects"].append(
            {
                "id": project["id"],
                "name": project["name"],
                "materials": list(materials.values()),
                "totalCost": total_cost,
                "itemCount": item_count,
            }
        )

    return jsonify(result)


# -------------------------
# EXPORT EXCEL
# -------------------------
@app.post("/api/export")
def export_excel():
    body = request.json
    project_id = body.get("projectId")

    projects = read_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "Project not found"}), 404

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Helper to add sheet
    def add_sheet(sheet_name, sheet_data, headers):
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)

        for row in sheet_data:
            ws.append([row.get(field, "") for field in headers])

    # Export new sheets
    if "sheets" in project:
        for sheet_name, sheet in project["sheets"].items():
            headers = sheet.get("headers", [])
            data_rows = sheet.get("data", [])
            add_sheet(sheet_name, data_rows, headers)

    # Create file
    filename = f"export_{project['name']}_{project['id']}.xlsx"
    path = os.path.join(DATA_DIR, filename)
    wb.save(path)

    return send_file(path, as_attachment=True)


# -------------------------
# RUN
# -------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=PORT)
